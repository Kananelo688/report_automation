#*******************************************************************************************
# Description: Module that define function to automate Packet Loss Report Generation
#
#Filename: packet_loss.py
#Author: Kananelo Chabeli
#*******************************************************************************************


import data_processing as dp
import openpyxl as xl
from openpyxl.styles import PatternFill, Font
from pathlib import Path
ROUTER_LOOKUP = {"ABIA":"ABIA-6804_1",
"AIRPORT":"MZD_6804E_1",
"ALWYNSKOP":"QCG-6804_1",
"AME":"HQ-M6000-8S",
"BAFALI":"MQQ-6804_1",
"BAKING":"MPU-6804_1",
"BATI":"NYA-6804_1",
"BEREA PLATEAU":"BEP-6804-1",
"BEREA_HOSPITAL":"TYG-6804_1",
"BEREAPLATEAU":"BEP-6804-1",
"BESELE":"ABIA-6804_1",
"BOBATSI":"MQQ-6804_1",
"BOCHABELA":"MBT-6804_1",
"BOINYATSO":"MBT-6804_1",
"BOROKHOANENG":"BRG-6804_1",
"BOSOFO":"ABIA-6804_1",
"BOTSOLA":"LRB-6804_1",
"BRAAKFONTEIN":"MOK-6804_1",
"BULARA":"BHB-6804_1",
"BUOASONO":"MBT-6804_1",
"BUTHA BUTHE":"BHB-6804_1",
"BUTHABUTHE":"BHB-6804_1",
"BUTHABUTHE_TLOKOENG":"BHB-6804_1",
"CALEDON":"BHB-6804_1",
"CARLTONCENTRE":"HQ-M6000-8S",
"CHAFO":"LRB-6804_1",
"CORNEXCHANGE":"MPU-6804_1",
"DR_GUESTHOUSE":"MBT-6804_1",
"ETLHQ":"MSU-8905E-CORE-SW01",
"FLORIDA":"HQ-M6000-8S",
"FOKOTHI":"HQ-M6000-8S",
"FOSO":"MBT-6804_1",
"FOSO_2":"MBT-6804_1",
"GOLDEN_TOWER":"MAE-6804_1",
"HA KOATSI":"NYA-6804_1",
"HA MPITI":"QSK-6804_1",
"HA NYANE":"NYA-6804_1",
"HA QABA":"MOK-6804_1",
"HA RAMABANTA":"MOK-6804_1",
"HA SEHAPA":"QSK-6804_1",
"HA SEHLOHO":"QSK-6804_1",
"HA SETHO":"THE-M6000-8S",
"HA SHOAEPANE":"MRJ-6804_1",
"HA TSEPO":"MOK-6804_1",
"HA_MPUTANA":"NYA-6804_1",
"HA_NTINA":"MPU-6804_1",
"HA_NTOPO":"NYA-6804_1",
"HA_NTSELI":"TTA-6804_1",
"HA_SELESO":"ABIA-6804_1",
"HAKAMOHO":"BHB-6804_1",
"HAKEPI":"TYG-6804_1",
"HAKOATSI":"NYA-6804_1",
"HALEKOALA":"MZD_6804E_1",
"HALEPOLESA":"MAE-6804_1",
"HAMAEPHO":"NYA-6804_1",
"HAMOKHALINYANE":"MRJ-6804_1",
"HAMOPELI":"BHB-6804_1",
"HAMOTAU":"NYA-6804_1",
"HAMOTSOANE":"BHB-6804_1",
"HANTHO":"QCG-6804_1",
"HA-NTHO":"QCG-6804_1",
"HATHAMAE":"BRG-6804_1",
"HATLALI":"BEP-6804-1",
"HATLELAI":"ABIA-6804_1",
"HATSEPO":"MOK-6804_1",
"HATSIU":"MBT-6804_1",
"HIGHVELT":"MBT-6804_1",
"HILLSVIEW":"HQ-M6000-8S",
"HLALELE":"THE-M6000-8S",
"HLEOHENG":"MPU-6804_1",
"HOLYCROSS":"QCG-6804_1",
"INDUSTRIAL1":"HQ-M6000-8S",
"INDUSTRIAL2":"HQ-M6000-8S",
"JANTEU":"MQQ-6804_1",
"KAO":"BHB-6804_1",
"KAO_MINE":"BHB-6804_1",
"KAONYANE":"MAHLASELA (L3 Switch)",
"KATSE":"TTA-6804_1",
"KETANE":"QCG-6804_1",
"KETANE_TIPING":"NYA-6804_1",
"KHAFUNG":"TYG-6804_1",
"KHANYANE":"LRB-6804_1",
"KHATIBE_VALLEY":"BHB-6804_1",
"KHOSHANENG":"TTA-6804_1",
"KHUBELU":"ABIA-6804_1",
"KHUBELU_VALLEY":"MAPHOLANENG (L3 Switch)",
"KHUBETSOANA_DLM":"MBT-6804_1",
"KHUBETSOANA_LECOOP":"MBT-6804_1",
"KHUBETSOANA_WILLIS":"MBT-6804_1",
"KHUBETSOANA1":"MBT-6804_1",
"KHUBETSOANA2":"MBT-6804_1",
"KHUBETSOANADLM":"MBT-6804_1",
"KHUBETSOANAWILLIS":"MBT-6804_1",
"KOALABATA":"MBT-6804_1",
"KOALABATA_PRIMARY":"MBT-6804_1",
"KOLO":"BEP-6804-1",
"KOLONYAMA":"TYG-6804_1",
"KOTA":"LRB-6804_1",
"KOTISPHOLA":"MQQ-6804_1",
"KUBAKE":"MOK-6804_1",
"KUBUNG":"QCG-6804_1",
"LAKESIDE":"HQ-M6000-8S",
"LCE":"HQ-M6000-8S",
"LEBELONYANE":"QCG-6804_1",
"LEJONE":"TTA-6804_1",
"LEKOKOANENG":"TYG-6804_1",
"LELOALENG":"NYA-6804_1",
"LEQELE":"BRG-6804_1",
"LEQELE_2":"BRG-6804_1",
"LEQELE_BUSSTOP":"BRG-6804_1",
"LERIBE":"LRB-6804_1",
"LERIBE MORENENG":"BHB-6804_1",
"LERIBEMORENENG":"BHB-6804_1",
"LEROPONG":"QSK-6804_1",
"LESOBENG":"NYA-6804_1",
"LESOTHOHIGHSCHOOL":"HQ-M6000-8S",
"LESOTHOSUN":"HQ-M6000-8S",
"LETSENG":"LETS_6804_1",
"LEVISNEK":"BHB-6804_1",
"LIBONO":"BHB-6804_1",
"LIKHELIKA":"BHB-6804_1",
"LIKHOELE":"MAE-6804_1",
"LIKILENG":"BHB-6804_1",
"LINAKANENG":"TTA-6804_1",
"LINAKANENG_2":"TTA-6804_1",
"LIPHIRING":"MOK-6804_1",
"LIQALABENG":"QSK-6804_1",
"LIQHOBONG":"BHB-6804_1",
"LIQHOBONG_MINE":"TTA-6804_1",
"LIQOBONG":"BHB-6804_1",
"LISEMENG":"LRB-6804_1",
"LITHABANENG":"BRG-6804_1",
"LITHOTENG":"ABIA-6804_1",
"LOWER_SEOLI":"BRG-6804_1",
"LOWER_THAMAE":"BRG-6804_1",
"LOWERSEOLI":"BRG-6804_1",
"MABOLOKA":"MQQ-6804_1",
"MABOTE":"MBT-6804_1",
"MABOTEPOLICE":"MBT-6804_1",
"MACHACHE":"MZD_6804E_1",
"MAFETENG":"MAE-6804_1",
"MAFETENG EXCHANGE":"MAE-6804_1",
"MAFETENG_EXCHANGE":"MAE-6804_1",
"MAHLAKACHANENG":"QCG-6804_1",
"MAHLAKAPESE":"LRB-6804_1",
"MAHLASELA":"LETS_6804_1",
"MAHLOENYENG":"MRJ-6804_1",
"MAHLONG":"NYA-6804_1",
"MAHOBONG":"LRB-6804_1",
"MAJA":"MZD_6804E_1",
"MAJAHENG":"TYG-6804_1",
"MAJANE":"MRJ-6804_1",
"MAKHAKHE":"MRJ-6804_1",
"MAKHALENG":"MOK-6804_1",
"MAKHOATHI":"BRG-6804_1",
"MAKHOROANA":"MPU-6804_1",
"MAKOAE_1":"QCG-6804_1",
"MAKOAE2":"QCG-6804_1",
"MAKOANYANE":"BRG-6804_1",
"MAKOETJE":"MRJ-6804_1",
"MAKOPO":"BHB-6804_1",
"MALAOANENG":"LRB-6804_1",
"MALEALEA":"MOK-6804_1",
"MALEFILOANE":"MQQ-6804_1",
"MALEFILOANE REPEATER":"MQQ-6804_1",
"MALEHLAKANA":"TTA-6804_1",
"MALESAOANA":"LRB-6804_1",
"MALETSUNYANE":"NYA-6804_1",
"MALIMONG":"TYG-6804_1",
"MALINGOANENG":"MQQ-6804_1",
"MALUMENG":"MAE-6804_1",
"MAMAEBANA":"MOK-6804_1",
"MAMATHE":"TYG-6804_1",
"MANAMANENG":"TTA-6804_1",
"MANKOANENG":"LRB-6804_1",
"MANONYANE":"HQ-M6000-8S",
"MANTSEBO":"MZD_6804E_1",
"MANTSONYANE":"NYA-6804_1",
"MAPELENG":"MBT-6804_1",
"MAPETLA":"THE-M6000-8S",
"MAPHOHLOANE":"MOK-6804_1",
"MAPHOLANENG":"MQQ-6804_1",
"MAPOTENG":"MPU-6804_1",
"MAPOTENG_EXCHANGE":"MPU-6804_1",
"MAPUTSOE":"MPU-6804_1",
"MAPUTSOE_WASA":"MPU-6804_1",
"MAPUTSOEWASA":"MPU-6804_1",
"MAQELE":"MPU-6804_1",
"MAQHAKA":"MBT-6804_1",
"MAQOALA_VALLEY":"MOK-6804_1",
"MARABENG":"MBT-6804_1",
"MARAKABEI":"NYA-6804_1",
"MASANA":"MZD_6804E_1",
"MASERU WEST":"HQ-M6000-8S",
"MASERUMALL":"MSU-8905E-CORE-SW01",
"MASERUMALL_2":"HQ-M6000-8S",
"MASERUSUN":"HQ-M6000-8S",
"MASERUWEST":"HQ-M6000-8S",
"MASIANOKENG":"THE-M6000-8S",
"MASOWE":"THE-M6000-8S",
"MASOWE_1":"THE-M6000-8S",
"MATALA":"BRG-6804_1",
"MATALA_PHASE2":"BRG-6804_1",
"MATEBENG":"QSK-6804_1",
"MATELILE":"MRJ-6804_1",
"MATHATA":"MPU-6804_1",
"MATHATA2":"MPU-6804_1",
"MATHOLENG":"MAE-6804_1",
"MATHUOANENG":"MQQ-6804_1",
"MATLAKENG":"BHB-6804_1",
"MATLAMENG":"LRB-6804_1",
"MATSOAING":"MQQ-6804_1",
"MATSOKU":"TTA-6804_1",
"MATUKENG":"MPU-6804_1",
"MAZENOD":"MZD_6804E_1",
"MAZENOD_AIRPORT":"MZD_6804E_1",
"MERITING":"MOK-6804_1",
"MESITSANENG":"MOK-6804_1",
"METOLONG":"BEP-6804-1",
"MJANYANE":"QCG-6804_1",
"MOFOKA":"MZD_6804E_1",
"MOHALE":"NYA-6804_1",
"MOHALESHOEK EXCHANGE":"MOK-6804_1",
"MOHALESHOEKHILL":"MOK-6804_1",
"MOKHETHOANENG":"MBT-6804_1",
"MOKHOABONG":"QSK-6804_1",
"MOKHOKHONG":"MZD_6804E_1",
"MOKHOTLONG_EXCHANGE":"MQQ-6804_1",
"MOKHOTLONG_TOWN":"MQQ-6804_1",
"MOKHOTLONG_VALLEY":"MQQ-6804_1",
"MOKHOTLONGEXCHANGE":"MQQ-6804_1",
"MOKHOTLONGHILL":"MQQ-6804_1",
"MONONTSA":"BHB-6804_1",
"MOREMOHOLO":"MQQ-6804_1",
"MORIFI":"MOK-6804_1",
"MORIJA":"MRJ-6804_1",
"MOSALEMANE":"MPU-6804_1",
"MOSENEKE":"MOK-6804_1",
"MOSETOA":"TTA-6804_1",
"MOSHOESHOE_II":"HQ-M6000-8S",
"MOSHOESHOEII":"HQ-M6000-8S",
"MOTENG":"BHB-6804_1",
"MOTENG_2":"BHB-6804_1",
"MOTETE_SEKHOBE":"LETS_6804_1",
"MOTHAE_MINE":"LETS_6804_1",
"MOTIMPOSO":"MBT-6804_1",
"MOTLOHELOA":"MZD_6804E_1",
"MOTSEKUOA":"MRJ-6804_1",
"MOTSEMOCHA":"MBT-6804_1",
"MOTSEMOCHA2":"MBT-6804_1",
"MOTSOANE":"BHB-6804_1",
"MOUNT MOOROSI":"QCG-6804_1",
"MPHAKI":"QCG-6804_1",
"MPHARANE":"MOK-6804_1",
"NALELI":"MBT-6804_1",
"NGOAJANE":"BHB-6804_1",
"NHTC":"BRG-6804_1",
"NKA-O-BEE":"TTA-6804_1",
"NKOKAMELE":"MQQ-6804_1",
"NNELESE":"BRG-6804_1",
"NQECHANE":"BHB-6804_1",
"NTJABANE":"TYG-6804_1",
"NYAKOSOBA":"BEP-6804-1",
"NYENYE":"MPU-6804_1",
"NYENYE_2":"MPU-6804_1",
"PALACE":"TYG-6804_1",
"PANENG":"MQQ-6804_1",
"PARAY_HOSPITAL":"TTA-6804_1",
"PARLIAMENT":"HQ-M6000-8S",
"PATISING":"MQQ-6804_1",
"PEKA":"TYG-6804_1",
"PENAPENA":"THE-M6000-8S",
"PHAHAMENG":"MAE-6804_1",
"PHATLALLA":"MOK-6804_1",
"PITSENG":"LRB-6804_1",
"PITSENG_POLICE":"LRB-6804_1",
"POLIHALI":"MQQ-6804_1",
"PONOANE":"MZD_6804E_1",
"PONTSENG":"TTA-6804_1",
"POPA":"MQQ-6804_1",
"PULANE":"MZD_6804E_1",
"QACHA":"QSK-6804_1",
"QACHASNEK":"QSK-6804_1",
"QALAKHENG":"MOK-6804_1",
"QEME_LIKOTSI":"THE-M6000-8S",
"QHOALI_HA_KELEBONE":"QCG-6804_1",
"QHOASING":"QCG-6804_1",
"QHOLAQHOE":"BHB-6804_1",
"QOALING":"BRG-6804_1",
"QOALING_HAMAMPHO":"THE-M6000-8S",
"QOALINGHAMAMPHO":"THE-M6000-8S",
"QUTHING_LNBS":"QCG-6804_1",
"RADIOLESOTHO":"HQ-M6000-8S",
"RALEJOE":"MZD_6804E_1",
"RAMATSELISO":"QSK-6804_1",
"RAMATSHELISO":"QSK-6804_1",
"RAMOHAPI":"MAE-6804_1",
"RAMOKHELE":"MAE-6804_1",
"RAMOROBA":"QSK-6804_1",
"RAPOKOLANA":"NYA-6804_1",
"RAPOLEBOEA":"MOK-6804_1",
"RATJOMOSE":"THE-M6000-8S",
"ROMA_HA_SKAUT":"ROM-6804_1",
"ROMA_HASKAUT":"ROM-6804_1",
"ROMA_HATABUTLE":"ROM-6804_1",
"ROMAEXCHANGE":"ROM-6804_1",
"ROTHE":"MRJ-6804_1",
"ROYAL PALACE":"HQ-M6000-8S",
"ROYALPALACE":"HQ-M6000-8S",
"SANIBORDER":"MQQ-6804_1",
"SANQEBETHU":"MQQ-6804_1",
"SEAPOINT":"HQ-M6000-8S",
"SEATE":"MQQ-6804_1",
"SEBAPALA":"QCG-6804_1",
"SEBOCHE":"BHB-6804_1",
"SEBOTHOANE":"LRB-6804_1",
"SEFIKENG":"THE-M6000-8S",
"SEHLABATHEBE":"QSK-6804_1",
"SEHONGHONG":"MQQ-6804_1",
"SEHONGHONG":"TTA-6804_1",
"SEKAMANENG":"MBT-6804_1",
"SEKAMANENG_2":"MBT-6804_1",
"SELESO":"ABIA-6804_1",
"SEMONKONG":"NYA-6804_1",
"SEOKA":"NYA-6804_1",
"SEPHAPHOS GATE":"MAE-6804_1",
"SEROALANKHOANA":"TTA-6804_1",
"SERUTLE":"BHB-6804_1",
"SESHOTE":"TTA-6804_1",
"SETALA":"NYA-6804_1",
"SETIBING":"BEP-6804-1",
"SETLAKALLENG":"BHB-6804_1",
"SGSN":"MPBN_Switch",
"SHALABENG":"ABIA-6804_1",
"SHEPESELI":"BHB-6804_1",
"SILOE":"MAE-6804_1",
"SOLANE":"TTA-6804_1",
"SOURU":"QSK-6804_1",
"ST. JAMES":"MQQ-6804_1",
"ST_AGNES":"TYG-6804_1",
"ST_ROSE":"TYG-6804_1",
"ST_STEPHENS":"MOK-6804_1",
"STADIUM":"HQ-M6000-8S",
"STATE LIBRARY":"HQ-M6000-8S",
"STATELIBRARY":"HQ-M6000-8S",
"STJAMES":"MQQ-6804_1",
"TEYATEYANENG":"TYG-6804_1",
"THABA TSEKA_EXCHANGE":"TTA-6804_1",
"THABA TSEKA_HILL":"TTA-6804_1",
"THABA_PUTSOA":"BEP-6804-1",
"THABABOSIU":"BRG-6804_1",
"THABALIHLOLO":"TTA-6804_1",
"THABANA_MORENA":"MAE-6804_1",
"THABANENG":"MAE-6804_1",
"THABA-NTSHO":"MQQ-6804_1",
"THABA-TSEKA HILL":"TTA-6804_1",
"THAKABANNA":"BHB-6804_1",
"THETSANE_HOSPITAL":"THE-M6000-8S",
"THETSANE_INDUSTRIAL":"THE-M6000-8S",
"THETSANE_LEKHOOA":"THE-M6000-8S",
"THETSANE_LESIA":"THE-M6000-8S",
"THETSANEHOSPITAL":"THE-M6000-8S",
"THETSANEINDUSTRIAL":"THE-M6000-8S",
"THETSANELEKHOOA":"THE-M6000-8S",
"THETSANELESIA":"THE-M6000-8S",
"THETSANELESIA2":"THE-M6000-8S",
"THETSINYANE":"QCG-6804_1",
"THETSINYANE":"Lebolonyane (L3 Switch)",
"THUATHE":"BEP-6804-1",
"TIKOE":"THE-M6000-8S",
"TIPING":"MOK-6804_1",
"TLOKOENG":"MQQ-6804_1",
"TLOKOTSING":"MAE-6804_1",
"TOLOANE":"MRJ-6804_1",
"TSAKHOLO":"MAE-6804_1",
"TSEHLANYANE":"LRB-6804_1",
"TSIFALIMALI":"LRB-6804_1",
"TSIKOANE":"MPU-6804_1",
"TSIME":"BHB-6804_1",
"TSOLO":"THE-M6000-8S",
"TSOLO2":"THE-M6000-8S",
"TSOLOANE":"MOK-6804_1",
"TSOSANE":"MBT-6804_1",
"TSOSANE_PRIMARY":"MBT-6804_1",
"UPPER_MOYENI":"QCG-6804_1",
"UPPERMOYENI":"QCG-6804_1",
"VANROOYAN":"MAE-6804_1",
"VANROOYEN":"MAE-6804_1",
"VICTORIA":"HQ-M6000-8S",
"KHUBETSOANA_ECOL":"MBT-6804_1",
"BOKONG":"TTA-6804_1",
"MAKHANGOA":"TTA-6804_1",
"HA TLALI":"BEP-6804-1",
"THABA NTSO":"MQQ-6804_1",
"HA MOSITI":"LRB-6804_1"
}


def insert_data(kpi,routers,sheet=None,task='3G',conditional_format=50.00):
    """
    Inserts data of the given KPI into the given worksheet.

    ---------
    Parameters:
    sheet(openpyxl Worksheet): an object of Worksheet, which represent the sheet to which data will be written
    kpi(data_processing KPI): an object of KPI class, which has its data already initialized from the spread_sheets
    """
    dates = list()
    site_names = list()
    cell_font = Font(color = "FFB22222")
    bold_font = Font(bold= True)
    bold_color_font = Font(bold= True, color = "FFB22222")
    cell_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    number_format_4G = '0.00%'
    for key in kpi.data.keys():
        if key[0] not in dates:
            dates.append(key[0])
        if key[1] not in site_names:
            site_names.append(key[1])
            #print(key[1])
    
    sheet.cell(row = 1, column = 1).value = kpi.kpi_name
    sheet.cell(row = 1, column = 1).font = bold_font
    for date in dates:
        sheet.cell(row = 2, column = dates.index(date)+2).value = date
        sheet.cell(row = 2, column = dates.index(date)+2).font = bold_font
    router_data = dict()
    for router in routers:
        for date in dates:
            router_data[(date,router)] = list()
            for key in kpi.data:
                if '4G_' in key[1] or '3G_' in key[1]:
                    if key[0] == date and key[1][3:] in routers[router]:
                        router_data[(date,router)].append(kpi.data[key])
                else:
                    if key[0] == date and key[1] in routers[router]:
                        router_data[(date,router)].append(kpi.data[key])
    dump_kpi = dp.KPI('dump',kpi.agg_fun)
    for key in list(router_data.keys()):
        if len(router_data[key]) == 0:
            del router_data[key]
    dump_kpi.generate_pivot_table(router_data)
    router_data = dump_kpi.data
    #Now have to write the data to the external sheet
    row_num = 3
    move_flag = False
    for router in routers:
        sheet.cell(row = row_num, column = 1).value = router
        sheet.cell(row = row_num, column = 1).font = bold_font
        for date in dates:
            if (date,router) in router_data:
                move_flag = False
                sheet.cell(row = row_num, column = dates.index(date)+2).value = router_data[(date,router)]
                sheet.cell(row = row_num, column = dates.index(date)+2).font = bold_font
                if router_data[(date,router)]>conditional_format:
                    sheet.cell(row = row_num, column = dates.index(date)+2).font = bold_color_font
                    sheet.cell(row = row_num, column = dates.index(date)+2).fill = cell_fill
                if task == '4G':
                    sheet.cell(row = row_num, column = dates.index(date)+2).number_format = number_format_4G
            else:
                move_flag = True
                break
                
            
        
        for site in routers[router]:
            if move_flag:
                break
            row_num +=1
            if task+'_'+site in site_names:
                site = task+'_'+site
            elif site in site_names:
                pass
            else:
                row_num-=1
                continue
            sheet.cell(row = row_num,column = 1).value = site
            
            for date in dates:
                if (date,site) in kpi.data:
                    sheet.cell(row = row_num,column = dates.index(date)+2).value = kpi.data[(date,site)]
                    if kpi.data[(date,site)]>conditional_format:
                        sheet.cell(row = row_num,column = dates.index(date)+2).font = cell_font
                        sheet.cell(row = row_num, column = dates.index(date)+2).fill= cell_fill
                    if task == '4G':
                        sheet.cell(row = row_num,column = dates.index(date)+2).number_format = number_format_4G
                else:
                    sheet.cell(row = row_num,column = dates.index(date)+2).value = None
                    
        row_num+=1

def compile_report(destination_filename = 'Packetloss.xlsx',verbose = False):
    """
    Computes Packet Loss Report, The filenames of the stats for 4G and 3G must exists in the directory, with specific naming:
        1. 4G_Packetloss.xlsx for 4G Stats
        2. 3G_Packetloss.xlsx for 3G Stats

    ---------
    Parameters:
    filename_4G_stats(str): File name of the combined 4G Statistics exported from the management(combined SDR and ITBBU).
    
    filename_3G_stats(str): Filename of the  3G Daily Packet Info stats.
    
    destination_filename(str): Filename of the destination report. Defaults to 'Packetloss'
    
    verbose(Bool): Flag used for debugging, if set to True, print the progress of report compilation.
    """
    if verbose:
        print('Analysing files in the working directory...', end = '', flush = True)
    filename_3G_stats = None
    filename_4G_stats = None

    files = Path('.').iterdir()

    for file in files:
        if file.name == '4G_Packetloss.xlsx':
            filename_4G_stats = '4G_Packetloss.xlsx'
        if file.name == '3G_Packetloss.xlsx':
            filename_3G_stats = '3G_Packetloss.xlsx'
    if filename_3G_stats is None:
        raise ValueError('3G Packetloss Stats not found. Please make that it exists and is named: "3G_Packetloss.xlsx"')
    if filename_4G_stats   is None:
        raise ValueError('4G Packetloss Stats not found. Please make sure it exists and is named: "4G_Packetloss.xlsx"')
    
    if verbose:
        print('done\nAll filed Found. Load files...',end= '', flush = True)
    
    wb_4G = xl.load_workbook(filename_4G_stats)
    wb_3G = xl.load_workbook(filename_3G_stats)
    
    if verbose:
        print('done.\nCreating Report workbook and Adding Sheets...',end = '' , flush = True)
    
    report_wb = xl.Workbook()
    report_wb.remove(report_wb['Sheet'])
    report_wb.create_sheet('3G PACKET LOSS')
    report_wb.create_sheet('3G PACKET DELAY')
    report_wb.create_sheet('3G PACKET JITTER')
    report_wb.create_sheet('VoLTE LOSS')
    
    if verbose:
        print('done.\nCreating KPIs...',end = '', flush = True)
    
    kpi_4G_volte_loss = dp.KPI('DL S1-U RTP Packet Loss Rate of QCI1(%)','mean')
    kpi_3G_packet_loss = dp.KPI('Rate of Lost Detected-Packet(1/10000)','sum')
    kpi_3G_delay = dp.KPI('Mean End-to-End Delay of Detected Packe(ms)','max')
    kpi_3G_jitter = dp.KPI('Mean Jitter of Detected Packet(ms)','max')
    
    if verbose:
        print('done.\nPerforming Text to columns...',end = '', flush = True)
    
    dp.text_to_columns(wb_4G['Sheet0'],1)
    dp.text_to_columns(wb_3G['Sheet0'],1)
    
    if verbose:
        print('done.\nReading data from spreadsheets...',end = '' , flush = True)
    
    kpi_4G_volte_loss.generate_pivot_table(dp.pivot_table_data(wb_4G['Sheet0'],['Begin','ENBFunction Name'],kpi_4G_volte_loss.kpi_name))
    kpi_3G_packet_loss.generate_pivot_table(dp.pivot_table_data(wb_3G['Sheet0'],['Begin','Office Name'],kpi_3G_packet_loss.kpi_name))
    kpi_3G_delay.generate_pivot_table(dp.pivot_table_data(wb_3G['Sheet0'],['Begin','Office Name'],kpi_3G_delay.kpi_name))
    kpi_3G_jitter.generate_pivot_table(dp.pivot_table_data(wb_3G['Sheet0'],['Begin','Office Name'],kpi_3G_jitter.kpi_name))
    
    routers = dict()
    for key in ROUTER_LOOKUP:
        if ROUTER_LOOKUP[key] in routers:
            routers[ROUTER_LOOKUP[key]].append(key)
        else:
            routers[ROUTER_LOOKUP[key]]  = [key]
    if verbose:
        print('done.\nCompiling 3G Packet Loss...', end = '' , flush = True)
    insert_data(kpi_3G_packet_loss,routers,sheet = report_wb['3G PACKET LOSS'], task = '3G', conditional_format=50.00)
    if verbose:
        print('done.\nCompiling 3G Packet Delay...', end = '', flush = True)
    insert_data(kpi_3G_delay,routers,sheet = report_wb['3G PACKET DELAY'], task = '3G', conditional_format=5)
    if verbose:
        print('done.\nCompiling 3G Packet Jitter...', end = '', flush = True)
    insert_data(kpi_3G_jitter,routers,sheet = report_wb['3G PACKET JITTER'], task = '3G', conditional_format=5)
    if verbose:
        print('done.\nCompiling LTE loss...', end = '', flush = True)
    insert_data(kpi_4G_volte_loss,routers,sheet = report_wb['VoLTE LOSS'], task = '4G', conditional_format=0.0005)
    if verbose:
        print('done.\nSaving results...',end = '', flush = True)
    report_wb.save(destination_filename)
    wb_3G .close()
    wb_4G.close()
    report_wb.close()
    if verbose:
        print('done',flush = True,)


# In[32]:


def main(verbose=True):
    compile_report(verbose=True)

if __name__ == 'main':
    main()



