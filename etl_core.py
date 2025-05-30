#**********************************************************************************
# Description: Module that defines Function that automates computation ETL Core 
#               data of the Qlik Report
# Filename: etl_core.py
# Author: Kananelo Chabeli
#*********************************************************************************

import openpyxl as xl
from datetime import datetime,timedelta,date
import data_processing as dp
from pathlib import Path


# for file in Path('.').iterdir():
#     if file.is_file():
#         print(file.name)

def insert_data(worksheet, row_idx,work,num_inserts = 7,**kwargs):
    """
    Inserts computed data into the worksheet, given the KPIs information
    """
    if work in ['2G','3G']:
        core_PDP_kpi = kwargs['core_PDP']
        core_SAU_kpi = kwargs['core_SAU']
        dr_PDP_kpi = kwargs['dr_PDP']
        dr_SAU_kpi = kwargs['dr_SAU']
        
        for i in range(1,num_inserts+1):
            insert_date = worksheet.cell(row = row_idx+i,column=3).value.strftime("%Y-%m-%d") #get the start date where to insert
            
            core_PDP = core_PDP_kpi.data[insert_date]
            core_SAU = core_SAU_kpi.data[insert_date]
            dr_PDP = dr_PDP_kpi.data[insert_date]
            dr_SAU = dr_SAU_kpi.data[insert_date]
            #print(f'Date:{insert_date}, core_PDP:{core_PDP},core_SAU: {core_SAU},dr_PDP:{dr_PDP},dr_SAU:{dr_SAU}')
            worksheet.cell(row = row_idx+i,column=4).value = (core_PDP+dr_PDP)/8
            worksheet.cell(row = row_idx+i,column=5).value = (core_SAU+dr_SAU)/8
            #COPY CELLS FORMAST
            dp.copy_cell_format(worksheet.cell(row = row_idx,column=4),worksheet.cell(row = row_idx+i,column=4))
            dp.copy_cell_format(worksheet.cell(row = row_idx,column=5),worksheet.cell(row = row_idx+i,column=5))
    elif work == 'VLR':
        core_online_subs = kwargs['core_online_subs']
        dr_online_subs = kwargs['dr_online_subs']
        core_num_subs = kwargs['core_num_subs']
        dr_num_subs = kwargs['dr_num_subs']

        for i in range(1,num_inserts+1):
            insert_date = worksheet.cell(row = row_idx+i,column=3).value.strftime("%Y-%m-%d") #get the start date where to insert
            core_online = core_online_subs.data[insert_date]
            dr_online = dr_online_subs.data[insert_date]

            core_num = core_num_subs.data[insert_date]
            dr_num = dr_num_subs.data[insert_date]

            worksheet.cell(row = row_idx+i,column=7).value = (dr_num+core_num)/8
            worksheet.cell(row = row_idx+i,column=8).value = (dr_online+core_online)/8

            dp.copy_cell_format(worksheet.cell(row = row_idx,column=7),worksheet.cell(row = row_idx+i,column=7))
            dp.copy_cell_format(worksheet.cell(row = row_idx,column=8),worksheet.cell(row = row_idx+i,column=8))
    elif work == '4G':
        dr_EPS_attach = kwargs['dr_EPS_attach']
        dr_NSA_attach = kwargs['dr_NSA_attach']
        core_EPS_attach = kwargs['core_EPS_attach']

        dr_default_bearers = kwargs['dr_default_bearers']
        core_default_bearers = kwargs['core_default_bearers']

        for i in range(1, num_inserts+1):
            insert_date = worksheet.cell(row = row_idx+i,column=3).value.strftime("%Y-%m-%d") #get the start date where to insert

            attach = dr_EPS_attach.data[insert_date]+dr_NSA_attach.data[insert_date]+core_EPS_attach.data[insert_date]

            bearers = dr_default_bearers.data[insert_date] + core_default_bearers.data[insert_date]
            
            worksheet.cell(row = row_idx+i,column=5).value = attach/8
            worksheet.cell(row = row_idx+i,column=4).value = bearers/8
            
            dp.copy_cell_format(worksheet.cell(row = row_idx,column=4),worksheet.cell(row = row_idx+i,column=4))
            dp.copy_cell_format(worksheet.cell(row = row_idx,column=5),worksheet.cell(row = row_idx+i,column=5))
    else:
        raise ValueError(f"Invalid 'work' given. Allowed values are '2G','3G','4G',pr 'VLR'. Given {work}")
       

def etl_core(end_date,verbose=False):
    """"
        Central processor all the ETL Core Sheet. This script read filenames from the current directory with following formats:
            1. All Network KPI Qlik 2019.xlsx: This is the filename of the Qlik file where data will be inserted.
            2. Core_Attach_KPIs.xlsx: This should be the filename of the attach KPIs exported from Core Network
            3. Core_uMACv_KPIs.xlsx: This is filename of the uMACV KPIS(LTE KPIs) from Core.
            4. Core_Vlr_KPIs.xlsx: This specifies the filename of the VLR
            5. Dr_Attach_KPIs.xlsx: This is filename of the attach KPIs from the DR.
            6. Dr_Vlr_KPIs.xlsx: This the filename of the Vlr KPIs from the DR
    """
    if end_date is None:
        raise ValueError('The end_date of Qlik where to insert data must be specified.')
    if verbose:
        print("Analysing files in the current directory...",end="",flush=True)
    
    files = Path('.').iterdir()
    qlik_filename = None
    core_attach_filename = None
    core_umac_filename = None
    core_vlr_filename = None
    dr_attach_filename = None
    dr_vlr_filename = None
    for file in files:
        if file.name == 'All Network KPI Qlik 2019.xlsx':
            qlik_filename = file.name
        elif file.name == 'Core_Attach_KPIs.xlsx':
            core_attach_filename = 'Core_Attach_KPIs.xlsx'
        elif file.name == 'Core_uMACv_KPIs.xlsx':
            core_umac_filename = 'Core_uMACv_KPIs.xlsx'
        elif file.name == 'Core_Vlr_KPIs.xlsx':
            core_vlr_filename = 'Core_Vlr_KPIs.xlsx'
        elif file.name == 'Dr_Attach_KPIs.xlsx':
            dr_attach_filename = 'Dr_Attach_KPIs.xlsx'
        elif file.name == 'Dr_Vlr_KPIs.xlsx':
            dr_vlr_filename = 'Dr_Vlr_KPIs.xlsx'
    if qlik_filename is None:
        raise ValueError('Qlik spreadsheet was not found. Please make sure that file "All Network KPI Qlik 2019.xlsx" exists.')
    if core_attach_filename is None:
        raise ValueError('Core Attach KPIs file not found. Please make sure the attach KPI spreadsheet from core exists and is named: "Core_Attach_KPIs.xlsx" .')
    if core_umac_filename is None:
        raise ValueError('uMACv Attach file not found. Please make sure that lTE attach(uMACv) spreadsheet from core exits and is named:"Core_uMACv_KPIs.xlsx".')
    if core_vlr_filename is None:
        raise ValueError('VLR file not found. Please make sure that the VLR User Measurement spreadsheet from the core exists and is named: "Core_Vlr_KPIs.xlsx".')
    if dr_attach_filename is None:
        raise ValueError ('Attach KPIs file from DR not found. Please make sure that the file exists and is  named: "Dr_Attach_KPIs.xlsx"')
    if dr_vlr_filename is None:
        raise ValueError('VLR User Measurement file from DR not found. Please make sure that the file exists and is named: "Dr_Vlr_KPIs.xlsx"')
    if verbose:
        print('done.\nAll Files found. Reading Spreadsheets...', end = ' ', flush = True)
    
    qlik_wb = xl.load_workbook(qlik_filename)
    core_attach_wb = xl.load_workbook(core_attach_filename)
    core_umac_wb = xl.load_workbook(core_umac_filename)
    core_vlr_wb = xl.load_workbook(core_vlr_filename)
    dr_vlr_wb = xl.load_workbook(dr_vlr_filename)
    dr_attach_wb = xl.load_workbook(dr_attach_filename)
    
    if verbose:
        print("done.\nReading data sheets...",end='',flush = True)
    
    qlik_sheet =qlik_wb['ETL Core']
    core_attach_sheet = core_attach_wb['sheet1']
    core_umac_sheet = core_umac_wb['sheet1']
    core_vlr_sheet = core_vlr_wb['sheet1']
    dr_attach_sheet = dr_attach_wb['Sheet0']
    dr_vlr_sheet = dr_vlr_wb['Sheet0']
    
    if verbose:
        print("done.\nPerforming text to columns on data...",end = '', flush = True)
    dp.text_to_columns(core_attach_sheet,2)
    dp.text_to_columns(core_umac_sheet,2)
    dp.text_to_columns(core_vlr_sheet,2)
    dp.text_to_columns(dr_attach_sheet,1)
    dp.text_to_columns(dr_vlr_sheet,1)
    
    if verbose:
        print("done.\nInitializing ETL Core sheet...",end = '', flush = True)
    
    dp.insert_rows(qlik_sheet,end_date)
    
    if verbose:
        print("done.\nInitializing ETL Core KPIs...",end = '', flush = True)
    #2G PDP KPI from CORE
    kpi_core_2G_PDP = dp.KPI('Maximum activated PDP contexts-GSM','max')
    #3G PDP KPI FROM CORE
    kpi_core_3G_PDP = dp.KPI('Maximum activated PDP contexts-UMTS','max')
    #2G PDP KPI FROM DR
    kpi_dr_2G_PDP =dp.KPI('Maximum Number of activated PDP contexts(GSM)','max')
    #3G PDP KPI FROM DR
    kpi_dr_3G_PDP = dp.KPI('Maximum Number of activated PDP contexts(UMTS)','max')

    #2G ATTACH KPI FROM DR
    kpi_dr_2G_SAU = dp.KPI('Maximum number of attached subscribers(GSM)','max')

    #3G ATTACH KPI FROM DR
    kpi_dr_3G_SAU = dp.KPI('Maximum number of attached subscribers(UMTS)','max')
    
    #2G ATTACH KPI FROM CORE
    kpi_core_2G_SAU = dp.KPI('Maximum number of attached subscribers(GSM)','max')
    #3G ATTACH KPI FROM CORE
    kpi_core_3G_SAU = dp.KPI('Maximum number of attached subscribers(UMTS)','max')
    
    #LTE ATTACH KPI FROM DR
    kpi_dr_EPS_attach = dp.KPI('Max Number of EPS Attach subscribers','max')
    #NSA(5G) ATTACH KPI FROM DR
    kpi_dr_NSA_attach = dp.KPI('Max Number of EPS Attach NSA subscribers','max')
     #LTE ATTACH FROM CORE
    kpi_core_EPS_attach = dp.KPI('Max Number of EPS Attach subscribers','max')
   
    #LTE PDP FROM DR
    kpi_dr_defualt_bearers = dp.KPI('Mean number of default bearers in active state','max')
    #LTE PDP FROM CORE
    kpi_core_defualt_bearers = dp.KPI('Mean number of default bearers in active state','max')
   
    
    #LTE
    kpi_core_vlr_num_subs = dp.KPI('Number of Subscribers in VLR','max')
    kpi_core_vlr_online_subs = dp.KPI('Number of On-line Subscribers in VLR','max')
    kpi_dr_vlr_num_subs = dp.KPI ('Number of Subscribers in VLR','max')
    kpi_dr_vlr_online_subs = dp.KPI('Number of On-line Subscribers in VLR','max')
    if verbose:
        print("done.\nInitializing KPI data(pivot tables)...",end='',flush=True)
    
    #Generating data for 2G and 3G PDP KPI
    kpi_core_2G_PDP.generate_pivot_table(dp.pivot_table_data(core_attach_sheet,['Start'],kpi_core_2G_PDP.kpi_name))
    kpi_core_3G_PDP.generate_pivot_table(dp.pivot_table_data(core_attach_sheet,['Start'],kpi_core_3G_PDP.kpi_name))
    kpi_dr_2G_PDP.generate_pivot_table(dp.pivot_table_data(dr_attach_sheet,['Begin'],kpi_dr_2G_PDP.kpi_name))
    kpi_dr_3G_PDP.generate_pivot_table(dp.pivot_table_data(dr_attach_sheet,['Begin'],kpi_dr_3G_PDP.kpi_name))

    #Generating data for 2G and 3G Attach KPI
    kpi_core_2G_SAU.generate_pivot_table(dp.pivot_table_data(core_attach_sheet,['Start'],kpi_core_2G_SAU.kpi_name))
    kpi_core_3G_SAU.generate_pivot_table(dp.pivot_table_data(core_attach_sheet,['Start'],kpi_core_3G_SAU.kpi_name))
    kpi_dr_2G_SAU.generate_pivot_table(dp.pivot_table_data(dr_attach_sheet,['Begin'],kpi_dr_2G_SAU.kpi_name))
    kpi_dr_3G_SAU.generate_pivot_table(dp.pivot_table_data(dr_attach_sheet,['Begin'],kpi_dr_3G_SAU.kpi_name))
    
    #Generating data for LTE Attach KPIs
    kpi_dr_EPS_attach.generate_pivot_table(dp.pivot_table_data(dr_attach_sheet,['Begin'],kpi_dr_EPS_attach.kpi_name))
    kpi_dr_NSA_attach.generate_pivot_table(dp.pivot_table_data(dr_attach_sheet, ['Begin'],kpi_dr_NSA_attach.kpi_name))
    kpi_core_EPS_attach.generate_pivot_table(dp.pivot_table_data(core_umac_sheet,['Start'],kpi_core_EPS_attach.kpi_name))

    #Generating data for LTE PDP KPI
    kpi_dr_defualt_bearers.generate_pivot_table(dp.pivot_table_data(dr_attach_sheet,['Begin'],kpi_dr_defualt_bearers.kpi_name))
    kpi_core_defualt_bearers.generate_pivot_table(dp.pivot_table_data(core_umac_sheet,['Start'],kpi_core_defualt_bearers.kpi_name))

    #Generating data for VLR subscribers
    kpi_core_vlr_num_subs.generate_pivot_table(dp.pivot_table_data(core_vlr_sheet,['Start'],kpi_core_vlr_num_subs.kpi_name))
    kpi_core_vlr_online_subs.generate_pivot_table(dp.pivot_table_data(core_vlr_sheet,['Start'],kpi_core_vlr_online_subs.kpi_name))
    kpi_dr_vlr_num_subs.generate_pivot_table(dp.pivot_table_data(dr_vlr_sheet,['Begin'],kpi_dr_vlr_num_subs.kpi_name))
    kpi_dr_vlr_online_subs.generate_pivot_table(dp.pivot_table_data(dr_vlr_sheet,['Begin'], kpi_dr_vlr_online_subs.kpi_name))
    
    if verbose:
        print("done.\nComputing 2G Attach a nd PDP data...",end = '',flush = True)
   
    indx = dp.search_insert(qlik_sheet, end_date)
    insert_data(qlik_sheet,indx[0],'2G',
                core_PDP = kpi_core_2G_PDP,
                core_SAU = kpi_core_2G_SAU, 
                dr_PDP = kpi_dr_2G_PDP, 
                dr_SAU =kpi_dr_2G_SAU)
    if verbose:
        print("done.\nComputing 3G Attach and PDP Contexts data...",end = '', flush = True)
    insert_data(qlik_sheet,indx[1],'3G',
                core_PDP = kpi_core_3G_PDP,
                core_SAU = kpi_core_3G_SAU, 
                dr_PDP = kpi_dr_3G_PDP, 
                dr_SAU =kpi_dr_3G_SAU)
    if verbose:
        print("done.\nComputing LTE attach and PDP Contexts data ...",end = '', flush = True)

    insert_data(qlik_sheet,indx[2],'4G',
                dr_EPS_attach = kpi_dr_EPS_attach,
                dr_NSA_attach = kpi_dr_NSA_attach,
                core_EPS_attach =kpi_core_EPS_attach,
                dr_default_bearers = kpi_dr_defualt_bearers,
                core_default_bearers = kpi_core_defualt_bearers)
    if verbose:
        print('done.\nComputing VLR registered and VLR attached data ...', end = '', flush = True)
    insert_data(qlik_sheet,indx[0],'VLR',
                core_online_subs = kpi_core_vlr_online_subs,
                dr_online_subs = kpi_dr_vlr_online_subs,
                core_num_subs = kpi_core_vlr_num_subs,
                dr_num_subs =  kpi_dr_vlr_num_subs)
                
    if verbose:
        print('done.\nSaving files...',end = '',flush = True)
    qlik_wb.save('result_'+qlik_filename)
    qlik_wb.close()
    core_attach_wb.close()
    core_vlr_wb.close()
    dr_attach_wb.close()
    core_umac_wb.close()
    if verbose:
        print('done')
    
def main():
    date = datetime(2025,4,27).date()
    etl_core(date,verbose=True)
if __name__ == '__main__':
    main()
