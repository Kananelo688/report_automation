#!/usr/bin/env python
# 
# Data Processing: 
# A package that holds classes and function for manipulating data for automating ETL's report generating

import openpyxl as xl
import pandas as pd
from statistics import mean
from datetime import datetime,date,timedelta
from copy import copy


AGGREGATION_FUNCTIONS = ['mean','max','min','sum','count']

DISTRICT_LOOK_UP = {}

def pivot_table_data(worksheet,primary_keys:list,data_column:str)->dict:
    """
    Extracts the data that mimics spreadsheet's privot table.
    -------
    Parameters:
    worksheet(Worksheet): an openpyxl worksheet object.

    primary_keys(list): list of columns names that will be used as primary keys. Must be of max of 2(row,col).

    data_column(str):The column number of the data field for this table.
    """
    #validate arguments
    simple= True
    if len(primary_keys)>2:
        raise ValueError(f"Expected at most 2 primary_keys. Given {len(primary_keys)} primary keys")
    elif len(primary_keys)==2:
        if primary_keys[0] == primary_keys[1]:
            simple = True
        else:
            simple = False
            
    data = dict()

    #search for the given column names:
    cols = worksheet.iter_cols( max_row = 1,values_only=True)
    col_names = list()
    for col in cols:
        col_names.append(col[0])

    
    if data_column in col_names:
        data_column_number = col_names.index(data_column)+1
    else:
        raise ValueError(f"Given data_column  of '{data_column}' does not exist in the sheet.")
    if primary_keys[0] in col_names:
        key_1 = col_names.index(primary_keys[0])+1
    else:
        raise ValueError(f"Given column name of '{primary_keys[0]}' does not exist in the sheet.")
    if not simple:
        if primary_keys[1] in col_names:
            key_2 = col_names.index(primary_keys[1])+1
        else:
            raise ValueError(f"Given column name of '{primary_keys[1]}' does not exist in the sheet.")
            
        
    for row in worksheet.iter_rows(min_row = 2,min_col = 1, max_col = data_column_number):
        if simple:
            key = row[key_1-1].value
        else:
            key = (row[key_1-1].value,row[key_2-1].value)
        if isinstance(key,datetime):
            key = key.date()
        if key in data:
            data[key].append(row[data_column_number-1].value)
        else:
            data[key] = [row[data_column_number-1].value]
    return data

def copy_cell_format(source_cell,target_cell):
    """
    Copies the format of cells 'source' into 'target' cell.
    """
    target_cell.font = copy(source_cell.font)
    target_cell.border = copy(source_cell.border)
    target_cell.fill = copy(source_cell.fill)
    target_cell.number_format = copy(source_cell.number_format)
    target_cell.protection = copy(source_cell.protection)
    target_cell.alignment = copy(source_cell.alignment)

def insert_rows(ws, last_date, num_rows=7):
    """
    Inserts `num_rows` rows after each occurrence of `last_date` in column 3.
    Copies columns 1–3, incrementing the date in column 3 by 1 day per new row.

    ------------
    Parameters:
        ws (Worksheet): openpyxl worksheet object.
        last_date (datetime.date): Date to search for in column 3.
        num_rows (int): Number of rows to insert after each match.
    """
    
    match_rows=search_insert(ws,last_date)

    # Process in reverse order to avoid affecting row indices
    for match_row in reversed(match_rows):
        insert_at = match_row + 1
        ws.insert_rows(insert_at, num_rows)

        for i in range(num_rows):
            for col in range(1, 4):  # Columns 1 to 3 only
                source_cell = ws.cell(row=match_row, column=col)
                target_cell = ws.cell(row=insert_at + i, column=col)

                # Copy value
                value = source_cell.value
                if col == 3 and hasattr(value, 'date'):
                    # Increment date for each new row
                    target_cell.value = value.date() + timedelta(days=i + 1)
                else:
                    target_cell.value = value

                # Copy formatting
                copy_cell_format(source_cell,target_cell)

def search_insert(ws,last_date):
    """
    Searches and returns the row numbers where insertion of new row will be made it ETL Core sheet.
    """
    match_rows = []
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=3).value
        if hasattr(cell_value, 'date'):
            if cell_value.date() == last_date:
                match_rows.append(row)
    if len(match_rows) == 0:
        raise ValueError(f'Given end date: {last_date} does not exist in the worksheet.')
    return match_rows

def text_to_columns(worksheet, column_number, separator=" "):
    """
    Split text into columns
    """
    for i in range(1, worksheet.max_row + 1):
        if isinstance(worksheet.cell(row=i, column= column_number).value,str):
            split_cell = worksheet.cell(row=i, column= column_number).value.split(separator)
            for j in range(len(split_cell)):
                worksheet.cell(row=i,column=column_number+j).value=split_cell[j]

class KPI:
    """
    Represents 3G or 4G KPI object.
    """
    def __init__(self,kpi_name, agg_fun):
        """
            Creates new instances of KPI object, and empty, data container.
        """
        self.kpi_name = kpi_name
        if agg_fun not in AGGREGATION_FUNCTIONS:
            raise ValueError(f"Given aggregation function of '{agg_fun}' is not supported. Supported functions are: {AGGREGATION_FUNCTIONS}.")
        self.agg_fun = agg_fun
        self.data = dict()
   
    def generate_pivot_table(self, data):
        """
            Adds the data of this KPI based on the agg_fun. The data parameter must be dictionary, generated
            by pivot_table_data function.
        """
        if self.agg_fun == 'mean':
            for key in data:
                self.data[key] = mean(data[key])
        elif self.agg_fun == 'max':
            for key in data:
                self.data[key] = max(data[key])
        elif self.agg_fun == 'min':
            for key in data:
                self.data[key] = min(data[key])
        elif self.agg_fun == 'sum':
            for key in data:
                self.data[key] = sum(data[key])
        elif self.agg_fun == 'count':
            for key in data:
                self.data[key] = len(data[key])


class Router:
    """
    Represents a Router to which several sites can be connected
    """
    def __init__(self, name):
        """
        Creates new instance of the Router.

        ---------
        Parameters:
        name(str): The name of the router
        """
        self.router_name = name
        self.sites = list()
    def add_site(self,site):
        """
        Adds new site into the list of this router's site, if it doesn't exist. If it exists, it raises ValueError
        """
        found = False
        for s in self.sites:
            if site.get_site_name() == s.get_site_name():
                found = True
                break
        if found:
            raise ValueError("Site Already Exists")
        else:
            self.sites.append(site)

class Site:
    """
    Represent a new Base Station.
    """
    def __init__(self, name=None):
        """
        Creates new instance of Site, with given name.
        """
        self.site_name = name
        self.router = None
        self.district = None
        self.KPIs = dict()
    
    def set_router(self, router):
        """
        sets the name of the rounter

        ----------
        Parameters:
        router(Router): The object of Router class, which represents the physical router to which this Siteis connected to.
        Raises  ArgumentError if the given argument is None or not the instance of Router class
        """
        if router is None or not isinstance(router, Router):
            raise ArgumentError("Invalid Argument")
        
        self.router = router
    
    def get_router(self)-> Router:
        """
        Returns the Router object, to which this Site is connected to.
        """
        return self.router
    
    def set_site_name(self,name):
        """
        Sets the name of this Site. Used to Change the name
        """
        if name is None:
            raise ArgumentError("Invalid Argument")
        
        self.site_name = name
    def get_site_name(self):
        """
        Returns the name of this Site
        """
        return self.site_name
    
    def set_district(self,district):
        """
        Sets the District where this Siteis found
        """
        if district is None:
            raise ArgumentError("Invalid Argument")
        self.district = district
    
    def get_district(self):
        """
        Returns the name of the district where this site is found
        """
        return self.district
    def add_KPI(self,kpi_name, kpi):
        """
        Adds a new KPI to the list of the Site's KPIs
        """
        if kpi_name in self.KPIs:
            raise KeyError("Given KPI already exists")
        else:
            self.KPIs[kpi_name]  = kpi




